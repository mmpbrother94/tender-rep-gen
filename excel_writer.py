from __future__ import annotations

import shutil
from copy import copy
from math import ceil, isclose
from pathlib import Path

from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font, PatternFill

from field_config import NOT_AVAILABLE
from synopsis_builder import FINAL_SYNOPSIS_LAYOUT


class ExcelWorkbookWriter:
    _CENTER_ALIGNMENT = Alignment(horizontal="center", vertical="center", wrap_text=True)
    _LEFT_ALIGNMENT = Alignment(horizontal="left", vertical="center", wrap_text=True)
    _DEFAULT_ROW_HEIGHT = 22.8
    _MAX_ROW_HEIGHT = 410.0
    _CATEGORY_FILLS = {
        101: "C6E0B4",
        102: "FFE699",
        103: "F4B183",
        104: "F8CBAD",
    }

    def write_outputs(
        self,
        extraction_bundle: dict[str, object],
        synopsis_template_path: str | Path,
        synopsis_output_path: str | Path,
        evaluation_bundle: dict[str, object],
        bid_template_path: str | Path,
        bid_output_path: str | Path,
    ) -> tuple[Path, Path]:
        synopsis_destination = self._copy_template(synopsis_template_path, synopsis_output_path)
        bid_destination = self._copy_template(bid_template_path, bid_output_path)

        self._write_synopsis_with_openpyxl(extraction_bundle, synopsis_destination)
        self._write_bid_evaluation_with_openpyxl(evaluation_bundle, bid_destination)
        return synopsis_destination, bid_destination

    def write_synopsis(self, extraction_bundle: dict[str, object], template_path: str | Path, output_path: str | Path) -> Path:
        destination = self._copy_template(template_path, output_path)
        self._write_synopsis_with_openpyxl(extraction_bundle, destination)
        return destination

    def write_bid_evaluation(self, evaluation_bundle: dict[str, object], template_path: str | Path, output_path: str | Path) -> Path:
        destination = self._copy_template(template_path, output_path)
        self._write_bid_evaluation_with_openpyxl(evaluation_bundle, destination)
        return destination

    def _copy_template(self, template_path: str | Path, output_path: str | Path) -> Path:
        source_template = Path(template_path).resolve()
        destination = Path(output_path).resolve()
        destination.parent.mkdir(parents=True, exist_ok=True)
        shutil.copy2(source_template, destination)
        return destination

    def _write_synopsis_with_openpyxl(self, extraction_bundle: dict[str, object], workbook_path: Path) -> None:
        workbook = load_workbook(workbook_path)
        worksheet = workbook.worksheets[0]
        self._prepare_synopsis_template_layout(worksheet)

        worksheet["A2"] = extraction_bundle.get("employer", NOT_AVAILABLE)
        worksheet["G1"] = f"Dt: {extraction_bundle.get('report_date', NOT_AVAILABLE)}"

        rows = extraction_bundle.get("synopsis_rows", extraction_bundle.get("rows", []))
        if isinstance(rows, list):
            self._write_synopsis_rows(worksheet, rows)

        self._mark_workbook_for_recalculation(workbook)
        workbook.save(workbook_path)

    def _write_bid_evaluation_with_openpyxl(self, evaluation_bundle: dict[str, object], workbook_path: Path) -> None:
        workbook = load_workbook(workbook_path)
        worksheet = workbook.worksheets[0]

        worksheet["I1"] = evaluation_bundle.get("report_date", NOT_AVAILABLE)
        worksheet["E20"] = evaluation_bundle.get("customer", NOT_AVAILABLE)
        worksheet["E21"] = evaluation_bundle.get("tender_ref", NOT_AVAILABLE)
        worksheet["E22"] = evaluation_bundle.get("work_title", NOT_AVAILABLE)
        worksheet["E23"] = evaluation_bundle.get("bid_submission_due_date", NOT_AVAILABLE)
        for row_number in range(20, 24):
            worksheet[f"E{row_number}"].alignment = copy(self._LEFT_ALIGNMENT)

        criteria = evaluation_bundle.get("criteria", [])
        if isinstance(criteria, list):
            criterion_rows = [
                int(criterion["row_number"])
                for criterion in criteria
                if isinstance(criterion, dict) and "row_number" in criterion
            ]
            total_criteria = len(criterion_rows)
            for index, criterion in enumerate(criteria):
                if not isinstance(criterion, dict):
                    continue
                row_number = int(criterion["row_number"])
                next_row = criterion_rows[index + 1] if index + 1 < total_criteria else 97
                option_rows = self._get_option_rows(worksheet, row_number, next_row)
                selected_row = self._find_selected_option_row(worksheet, option_rows, criterion)
                allocation_percent = float(
                    criterion.get("allocation_percent", float(criterion.get("allocation", 0.0)) * 100.0)
                )
                weighted_score = float(criterion.get("weighted_score", 0.0))

                for option_row in option_rows:
                    worksheet.cell(option_row, 7).value = ""
                    worksheet.cell(option_row, 7).number_format = "@"
                    worksheet.cell(option_row, 8).value = ""
                    worksheet.cell(option_row, 8).number_format = "0.00%"
                    worksheet.cell(option_row, 9).value = ""
                    worksheet.cell(option_row, 9).alignment = copy(self._LEFT_ALIGNMENT)

                worksheet.cell(selected_row, 7).value = self._format_percentage_text(allocation_percent)
                worksheet.cell(selected_row, 7).number_format = "@"
                worksheet.cell(selected_row, 8).value = weighted_score
                worksheet.cell(selected_row, 8).number_format = "0.00%"
                worksheet.cell(selected_row, 9).value = criterion.get("rationale", "")
                worksheet.cell(selected_row, 9).alignment = copy(self._LEFT_ALIGNMENT)

        worksheet["G97"] = "Actual Percentage"
        worksheet["H97"] = float(evaluation_bundle.get("total_fraction", 0.0))
        worksheet["H97"].number_format = "0.00%"
        total_percentage = float(evaluation_bundle.get("total_percentage", 0.0))
        worksheet["I97"] = (
            f"{evaluation_bundle.get('category', NOT_AVAILABLE)} - {evaluation_bundle.get('decision', NOT_AVAILABLE)} "
            f"(based on {total_percentage:.2f}%)"
        )
        worksheet["I97"].alignment = copy(self._LEFT_ALIGNMENT)
        self._highlight_category_row(worksheet, int(evaluation_bundle.get("category_row", 104)))

        self._mark_workbook_for_recalculation(workbook)
        workbook.save(workbook_path)

    def _prepare_synopsis_template_layout(self, worksheet: object) -> None:
        if not self._has_final_synopsis_layout(worksheet):
            worksheet.insert_rows(45, amount=6)
            self._copy_synopsis_row_styles(worksheet, template_row=44, target_rows=range(45, 51))

        merged_ranges = {str(range_ref) for range_ref in worksheet.merged_cells.ranges}
        if "A2:B2" not in merged_ranges:
            worksheet.merge_cells("A2:B2")
        if "G1:H1" not in merged_ranges:
            worksheet.merge_cells("G1:H1")

        worksheet["A2"] = "Name of Employer"
        worksheet["G1"] = "Dt: Report Date"
        for layout_row in FINAL_SYNOPSIS_LAYOUT:
            worksheet.cell(layout_row.row_number, 1).number_format = "@"
            worksheet.cell(layout_row.row_number, 2).number_format = "@"
            worksheet.cell(layout_row.row_number, 1).value = layout_row.serial
            worksheet.cell(layout_row.row_number, 2).value = layout_row.label

    def _copy_synopsis_row_styles(self, worksheet: object, template_row: int, target_rows: range) -> None:
        template_height = worksheet.row_dimensions[template_row].height
        max_column = max(worksheet.max_column, 7)
        for row_number in target_rows:
            if template_height is not None:
                worksheet.row_dimensions[row_number].height = template_height
            for column_number in range(1, max_column + 1):
                source_cell = worksheet.cell(template_row, column_number)
                target_cell = worksheet.cell(row_number, column_number)
                if source_cell.has_style:
                    target_cell._style = copy(source_cell._style)
                if source_cell.font:
                    target_cell.font = copy(source_cell.font)
                if source_cell.fill:
                    target_cell.fill = copy(source_cell.fill)
                if source_cell.border:
                    target_cell.border = copy(source_cell.border)
                if source_cell.alignment:
                    target_cell.alignment = copy(source_cell.alignment)
                if source_cell.protection:
                    target_cell.protection = copy(source_cell.protection)
                target_cell.number_format = source_cell.number_format

    def _has_final_synopsis_layout(self, worksheet: object) -> bool:
        current_label = str(worksheet.cell(41, 2).value or "").strip()
        return current_label == "Technology Provider"

    def _write_synopsis_rows(self, worksheet: object, rows: list[object]) -> None:
        row_map: dict[int, dict[str, object]] = {}
        for row_data in rows:
            if not isinstance(row_data, dict):
                continue
            row_number = int(row_data.get("row_number", 0))
            if row_number:
                row_map[row_number] = row_data

        start_row = FINAL_SYNOPSIS_LAYOUT[0].row_number
        end_row = FINAL_SYNOPSIS_LAYOUT[-1].row_number

        for row_number in range(start_row, end_row + 1):
            row_data = row_map.get(row_number, {})
            section = str(row_data.get("section", NOT_AVAILABLE))
            clause = str(row_data.get("clause", NOT_AVAILABLE))
            page = str(row_data.get("page", NOT_AVAILABLE))
            value = str(row_data.get("value", NOT_AVAILABLE))
            remark = str(row_data.get("remark", ""))

            centered_values = (section, clause, page)
            detail_values = (value, remark)

            for offset, cell_value in enumerate(centered_values, start=3):
                cell = worksheet.cell(row_number, offset)
                cell.value = cell_value
                cell.number_format = "@"
                cell.alignment = copy(self._CENTER_ALIGNMENT)

            for offset, cell_value in enumerate(detail_values, start=6):
                cell = worksheet.cell(row_number, offset)
                cell.value = cell_value
                cell.number_format = "@"
                cell.alignment = copy(self._LEFT_ALIGNMENT)

            worksheet.row_dimensions[row_number].height = self._estimate_synopsis_row_height(
                worksheet,
                row_number,
                centered_values + detail_values,
            )

    def _estimate_synopsis_row_height(
        self,
        worksheet: object,
        row_number: int,
        values: tuple[str, str, str, str, str],
    ) -> float:
        max_lines = 1
        for column_number, value in zip((3, 4, 5, 6, 7), values, strict=False):
            column_letter = worksheet.cell(1, column_number).column_letter
            width = worksheet.column_dimensions[column_letter].width or 12.0
            approx_chars_per_line = max(10, int(width * 1.1))
            wrapped_lines = 0
            text_value = value if value and value != NOT_AVAILABLE else ""
            for paragraph in str(text_value).splitlines() or [""]:
                wrapped_lines += max(1, ceil(len(paragraph) / approx_chars_per_line))
            max_lines = max(max_lines, wrapped_lines)

        estimated_height = max(self._DEFAULT_ROW_HEIGHT, (max_lines * 15.0) + 8.0)
        if row_number in (45, 46, 47, 48, 50) and estimated_height < 45.6:
            estimated_height = 45.6
        return min(self._MAX_ROW_HEIGHT, estimated_height)

    def _highlight_category_row(self, worksheet: object, category_row: int) -> None:
        for row_number in (101, 102, 103, 104):
            fill_color = self._CATEGORY_FILLS.get(row_number, "FFFFFF") if row_number == category_row else "FFFFFF"
            bold = row_number == category_row
            for cell in worksheet[f"B{row_number}:D{row_number}"][0]:
                updated_fill = PatternFill(fill_type="solid", fgColor=fill_color)
                updated_font = copy(cell.font) if cell.font else Font()
                updated_font.bold = bold
                cell.fill = updated_fill
                cell.font = updated_font

    def _get_option_rows(self, worksheet: object, start_row: int, next_start_row: int) -> list[int]:
        option_rows: list[int] = []
        for row_number in range(start_row, next_start_row):
            score = worksheet.cell(row_number, 6).value
            if isinstance(score, (int, float)):
                option_rows.append(row_number)
        return option_rows or [start_row]

    def _find_selected_option_row(self, worksheet: object, option_rows: list[int], criterion: dict[str, object]) -> int:
        allocation = float(criterion.get("allocation", 0.0))
        closest_row = option_rows[0]
        closest_gap = float("inf")
        for row_number in option_rows:
            score = worksheet.cell(row_number, 6).value
            if not isinstance(score, (int, float)):
                continue
            score_value = float(score)
            if isclose(score_value, allocation, rel_tol=1e-9, abs_tol=1e-9):
                return row_number
            gap = abs(score_value - allocation)
            if gap < closest_gap:
                closest_gap = gap
                closest_row = row_number
        return closest_row

    def _format_percentage_text(self, value: float) -> str:
        return f"{value:.2f}%"

    def _mark_workbook_for_recalculation(self, workbook: object) -> None:
        calculation = getattr(workbook, "calculation", None)
        if calculation is None:
            return
        if hasattr(calculation, "calcMode"):
            calculation.calcMode = "auto"
        if hasattr(calculation, "fullCalcOnLoad"):
            calculation.fullCalcOnLoad = True
        if hasattr(calculation, "forceFullCalc"):
            calculation.forceFullCalc = True


ExcelSynopsisWriter = ExcelWorkbookWriter
